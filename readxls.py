import os
import openpyxl
# Give the location of the file
my_path = "./CONSOLIDATED.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
my_max_col = my_sheet_obj.max_column
for i in range(1, my_max_col + 1):
   cell_obj = my_sheet_obj.cell(row = 9, column = i)
   print(cell_obj.value, end = " ")
